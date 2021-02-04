import openpyxl
import os
import pandas as pd
import numpy as np
from metrics import METRICS


def get_models(datasets):
    """It obtains the models used into the experiments"""
    dataframe = pd.read_csv('../results/' + datasets[0] + '/results.csv', sep=';')
    models = dataframe['MODEL'].unique()
    return models.tolist()


def get_best_prediction(results, metric, model, dataset):
    """It calculates the best prediction of one model in one dataset"""
    model_rows = results.loc[results['MODEL'] == model, :]
    best_value = 9999999999
    for index, row in model_rows.iterrows():
        path_y_test_denorm = '../data/' + dataset + '/' + np.str(row['NORMALIZATION']) + '/' + \
                             np.str(row['PAST_HISTORY_FACTOR']) + '/'

        path_preds = '../results/' + dataset + '/' + np.str(row['NORMALIZATION']) + '/' + np.str(
            row['PAST_HISTORY_FACTOR']) + '/' + np.str(row['EPOCHS']) + '/' + np.str(row['BATCH_SIZE']) + '/' + np.str(
            row['LEARNING_RATE']) + '/' + model + '/' + np.str(row['MODEL_INDEX']) + '.npy'

        y_test_denorm = np.load(path_y_test_denorm + 'y_test_denorm.np.npy').flatten()
        preds = np.load(path_preds).flatten()

        value = METRICS[metric](y_test_denorm, preds)

        if value < best_value:
            best_value = value
            best_model = row['MODEL_DESCRIPTION']

    return best_value, best_model


def create_excels():
    """It create the excels where the results are going to be saved"""
    if not os.path.exists('../results_best/'):
        os.mkdir('../results_best/')

    excel_metrics = pd.ExcelWriter('../results_best/metrics_by_predictions.xlsx', engine='openpyxl')
    excel_metrics.book = openpyxl.Workbook()

    excel_models = pd.ExcelWriter('../results_best/metrics_by_predictions_models.xlsx', engine='openpyxl')
    excel_models.book = openpyxl.Workbook()
    return excel_metrics, excel_models


def calculate_metrics(datasets, models, metrics, excel_metrics, excel_models):
    """It calculate the metrics, of each model in each dataset, and save them into the excel"""
    columns_names = ['dataset'] + models

    for metric in metrics:
        res_metric = pd.DataFrame(columns=columns_names).set_index('dataset')
        res_model = pd.DataFrame(columns=columns_names).set_index('dataset')
        for dataset in datasets:

            results = pd.read_csv('../results/' + dataset + '/results.csv', sep=';', index_col='Unnamed: 0')
            row_metric = []
            row_model = []

            for model in models:
                value, model_value = get_best_prediction(results, metric, model, dataset)
                row_metric.append(value)
                row_model.append(model_value)

            res_metric.loc[dataset, :] = row_metric
            res_model.loc[dataset, :] = row_model

        res_metric.to_excel(excel_metrics, sheet_name=metric)
        res_model.to_excel(excel_models, sheet_name=metric)

    return excel_metrics, excel_models


def save_excels(excel_metrics, excel_models):
    """It saves the excels with the information"""
    default_sheet_metrics = excel_metrics.book[excel_metrics.book.sheetnames[0]]
    excel_metrics.book.remove(default_sheet_metrics)
    excel_metrics.save()
    excel_metrics.close()

    default_sheet_models = excel_models.book[excel_models.book.sheetnames[0]]
    excel_models.book.remove(default_sheet_models)
    excel_models.save()
    excel_models.close()


def get_metrics():
    """Calculate the best values for a metrics of each model of each dataset, and saves the results into the sheets
    of an excel"""
    metrics = ['mse', 'rmse', 'mae', 'wape', 'mase']
    datasets = os.listdir('../results/')
    models = get_models(datasets)

    excel_metrics, excel_models = create_excels()
    excel_metrics, excel_models = calculate_metrics(datasets, models, metrics, excel_metrics, excel_models)
    save_excels(excel_metrics, excel_models)

    print('[INFO] Values of the metrics by predictions saved into "./results_best/metrics_by_predictions.xlsx"')
    print('[INFO] Models description of the best models saved into "./results_best/metrics_by_predictions_models.xlsx"')


if __name__ == '__main__':
    get_metrics()
